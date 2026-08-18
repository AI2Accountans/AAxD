import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Costeo Contrato 649 Davivienda"

# Mostrar líneas de cuadrícula
ws.views.sheetView[0].showGridLines = True

# Estilos
font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
font_regular = Font(name="Calibri", size=11, color="000000")
font_note = Font(name="Calibri", size=9, italic=True, color="595959")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fill_input = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo suave para campos editables

thin_border_side = Side(border_style="thin", color="D9D9D9")
thick_bottom_side = Side(border_style="medium", color="1F4E78")
double_bottom_side = Side(border_style="double", color="1F4E78")

border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
border_total = Border(top=thin_border_side, bottom=double_bottom_side)

# Título
ws.merge_cells("A1:F1")
ws["A1"] = "MODELO DE COSTEO BASADO EN ACTIVIDADES (ABC) - CONTRATO 649"
ws["A1"].font = font_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:F2")
ws["A2"] = "Análisis de Esfuerzo Técnico Estimado y Cotización de Soporte - Davivienda"
ws["A2"].font = font_subtitle
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

# Encabezados
headers = [
    "Categoría",
    "Módulos Impactados",
    "Esfuerzo Detallado (Según Correo)",
    "Horas / Mes (para cálculo)",
    "Valor Hora (COP)",
    "Subtotal Mensual (COP)"
]

ws.row_dimensions[4].height = 28
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_header

# Datos
data_rows = [
    [
        "Soporte Mensual",
        "GSKM_RETEFTE, GSKM_IVA, GSKM_RENTA, GSKM_FON_PERIODIC, GSKM_STEMMA",
        "8 a 10 Horas",
        10.0,
        0,
        "=D5*E5"
    ],
    [
        "Soporte Trimestral",
        "GSKM_FON, GSKM_SOC",
        "3 Horas (Amortizadas)",
        3.0,
        0,
        "=D6*E6"
    ],
    [
        "Soporte Anual / Ocasional",
        "GSKM_2516, GSKM_CARTERA CLIENTES, GSKM_ASOBANCARIA_MULTICASH",
        "2 Horas (Amortizadas)",
        2.0,
        0,
        "=D7*E7"
    ],
    [
        "Gestión Continua de IT",
        "Mantenimiento entorno Google Drive y licencias ALTOVA",
        "2 Horas",
        2.0,
        0,
        "=D8*E8"
    ]
]

for row_idx, row_data in enumerate(data_rows, 5):
    ws.row_dimensions[row_idx].height = 24
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = border_cell
        
        # Alineación y formato numérico
        if col_idx in [1, 3]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif col_idx == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "#,##0.0"
        elif col_idx == 5:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '"$"#,##0'
            cell.fill = fill_input # Resaltar en amarillo claro como editable
        elif col_idx == 6:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '"$"#,##0'

# Fila de Totales
total_row = 9
ws.row_dimensions[total_row].height = 26

cell_cat = ws.cell(row=total_row, column=1, value="TOTAL ESTIMADO")
cell_cat.font = font_bold
cell_cat.fill = fill_total
cell_cat.alignment = Alignment(horizontal="left", vertical="center")

cell_mod = ws.cell(row=total_row, column=2, value="10 Módulos")
cell_mod.font = font_bold
cell_mod.fill = fill_total
cell_mod.alignment = Alignment(horizontal="left", vertical="center")

cell_det = ws.cell(row=total_row, column=3, value="~17 Horas / Mes")
cell_det.font = font_bold
cell_det.fill = fill_total
cell_det.alignment = Alignment(horizontal="center", vertical="center")

cell_hrs = ws.cell(row=total_row, column=4, value="=SUM(D5:D8)")
cell_hrs.font = font_bold
cell_hrs.fill = fill_total
cell_hrs.alignment = Alignment(horizontal="center", vertical="center")
cell_hrs.number_format = "#,##0.0"

cell_vhr = ws.cell(row=total_row, column=5, value="Promedio / N.A.")
cell_vhr.font = font_bold
cell_vhr.fill = fill_total
cell_vhr.alignment = Alignment(horizontal="center", vertical="center")

cell_tot = ws.cell(row=total_row, column=6, value="=SUM(F5:F8)")
cell_tot.font = font_bold
cell_tot.fill = fill_total
cell_tot.alignment = Alignment(horizontal="right", vertical="center")
cell_tot.number_format = '"$"#,##0'

# Bordes de total
for col_idx in range(1, 7):
    ws.cell(row=total_row, column=col_idx).border = border_total

# Nota al pie
ws.merge_cells("A11:F11")
ws["A11"] = "* Nota: Diligencie los valores unitarios por hora en la columna 'Valor Hora (COP)' (resaltada en amarillo). Los subtotales y el total mensual se calcularán automáticamente."
ws["A11"].font = font_note

# Ajustar anchos de columnas
col_widths = {
    "A": 26,
    "B": 48,
    "C": 28,
    "D": 18,
    "E": 24,
    "F": 26
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

excel_path = r"C:\Users\IPHIX\Documents\Projects\DFRNT\BaseX\Modelo_Costeo_Soporte_Davivienda_Contrato_649.xlsx"
wb.save(excel_path)
print(f"Archivo Excel generado exitosamente en: {excel_path}")
